from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import os
import subprocess
import json
import os
import json
import time

# =====================================================
# LOAD CONFIG
# =====================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

config_path = os.path.join(BASE_DIR, "config.json")

with open(config_path, "r") as config_file:
    config = json.load(config_file)
    
EXCEL_FILE = config["excel_file"]
SHEET_NAME = config["sheet_name"]
START_ROW = config["start_row"]

URL_COLUMN = config["url_column"]
DIRECTORY_COLUMN = config["directory_column"]
FILENAME_COLUMN = config["filename_column"]

USER_DATA_DIR = config["user_data_dir"]
HEADLESS = config["headless"]

# =====================================================
# LOAD WORKBOOK
# =====================================================
wb = load_workbook(EXCEL_FILE)
ws = wb[SHEET_NAME]
records = []
row = START_ROW

while True:

    url_cell = ws[f"{URL_COLUMN}{row}"]
    dir_cell = ws[f"{DIRECTORY_COLUMN}{row}"]

    # Stop when URL column is blank
    if url_cell.value is None:
        break

    # Extract hyperlink
    if url_cell.hyperlink:

        url = url_cell.hyperlink.target

        # Directory from column L
        save_directory = str(dir_cell.value).strip()

        records.append({
            "row": row,
            "url": url,
            "directory": save_directory
        })

        print(f"Row {row}: hyperlink found")

    else:
        print(f"Row {row}: no hyperlink found")

    row += 1

print(f"\nTotal records found: {len(records)}")

# =====================================================
# PLAYWRIGHT
# =====================================================

with sync_playwright() as p:

    context = p.chromium.launch_persistent_context(
        user_data_dir=USER_DATA_DIR,
        headless=HEADLESS,
        accept_downloads=True
    )

    page = context.new_page()

    for index, record in enumerate(records, start=1):

        row_number = record["row"]
        url = record["url"]
        save_directory = record["directory"]

        print("\n================================================")
        print(f"Processing {index} of {len(records)}")

        try:

            # Create directory if needed
            os.makedirs(save_directory, exist_ok=True)

            # Open page
            page.goto(url, timeout=120000)

            # Wait for page load
            page.wait_for_load_state("networkidle")

            # Small delay for stability
            page.wait_for_timeout(3000)

            # =====================================================
            # DOWNLOAD
            # =====================================================

            with page.expect_download(timeout=120000) as download_info:

                page.get_by_role(
                    "link",
                    name="DOWNLOAD"
                ).click()

            download = download_info.value

            # Get filename
            filename = download.suggested_filename

            print(f"Downloading: {filename}")

            # Build full save path
            save_path = os.path.join(
                save_directory,
                filename
            )

            # Save file
            download.save_as(save_path)

            print(f"Saved to: {save_directory}")

            # =====================================================
            # WRITE FILENAME TO COLUMN O
            # =====================================================

            ws[f"{FILENAME_COLUMN}{row_number}"] = filename

            # Save workbook immediately
            wb.save(EXCEL_FILE)

            print(f"Updated Excel row {row_number}")

        except Exception as e:

            print("ERROR:")
            print(e)

# =====================================================
# FINAL SAVE
# =====================================================

wb.save(EXCEL_FILE)

print("\nAll downloads complete.")

input("\nPress ENTER to close...")