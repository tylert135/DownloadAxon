from openpyxl import load_workbook
from playwright.sync_api import sync_playwright
import os
import subprocess
import json
import sys
import time
from datetime import datetime
import argparse
import platform

# =====================================================
# GET ARGUMENTS (IF EXISTS)
# =====================================================
print(f"No. of args: {len(sys.argv)}")
parser = argparse.ArgumentParser()
parser.add_argument("--config_path")
args = parser.parse_args()

# =====================================================
# GET BASE DIRECTORY OF EXECUTABLE FOR PROJECT
# =====================================================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

print(f"Base directory: {BASE_DIR}")

# =====================================================
# LOAD CONFIG
# =====================================================
if args.config_path:
    print(f"Config path supplied: {args.config_path}{os.sep}config.json")
    config_path = os.path.join(args.config_path, "config.json")
else:
    print(f"Config path default: {BASE_DIR}{os.sep}config.json")
    config_path = os.path.join(BASE_DIR, "config.json")

with open(config_path, "r") as config_file:
    config = json.load(config_file)

EXCEL_FILE = config["excel_file"]

SHEET_NAME = config["sheet_name"]
START_ROW = config["start_row"]

URL_COLUMN = config["url_column"]
DIRECTORY_COLUMN = config["directory_column"]
FILENAME_COLUMN = config["filename_column"]

USER_DATA_DIR = config["user_data_dir"]
HEADLESS = config["headless"]

UNZIP_MOVE_BATCH = config["unzipMove_batch"]

CASE_STATUS = config["case_status"]

COMPLETED_COLUMN = config["completed_column"]

# =====================================================
# TIMER HELPER
# =====================================================

def format_elapsed(seconds):

    minutes, seconds = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)

    return f"{int(hours):02}:{int(minutes):02}:{seconds:05.2f}"

# =====================================================
# LOAD WORKBOOK
# =====================================================

if not os.path.exists(EXCEL_FILE):
    print(f"Excel file not found: {EXCEL_FILE}")
    input("\nPress ENTER to close...")
    exit()

print(f"Excel File: {EXCEL_FILE}")

wb = load_workbook(EXCEL_FILE)

ws = wb[SHEET_NAME]

records = []

row = START_ROW

while True:

    url_cell = ws[f"{URL_COLUMN}{row}"]

    dir_cell = ws[f"{DIRECTORY_COLUMN}{row}"]

    # Stop when URL column is blank
    if url_cell.value is None:
        break

    # Extract hyperlink
    if url_cell.hyperlink:

        url = url_cell.hyperlink.target

        # Normalize Windows path
        save_directory = os.path.normpath(
            str(dir_cell.value).strip()
        )

        records.append({
            "row": row,
            "url": url,
            "directory": save_directory
        })

        print(f"Row {row}: hyperlink found")
        print(f"Download file saved: {save_directory}")

    else:

        print(f"Row {row}: no hyperlink found")

    row += 1

print(f"\nTotal records found: {len(records)}")

chromium_dir = os.path.join(BASE_DIR, "browser", "chromium")

chromium_path = None

print(f"Chromiim Dir: {chromium_dir}")

if sys.platform.startswith("win"):
    for root, dirs, files in os.walk(chromium_dir):
        if "chrome.exe" in files:
            chromium_path = os.path.join(root, "chrome.exe")
            break
    if not chromium_path:
        raise FileNotFoundError("chrome.exe not found")
else:
    from pathlib import Path
    root = Path(f"{BASE_DIR}/browser/chromium")
    mac_dirs = [
        d for d in root.iterdir()
        if d.is_dir() and d.name.startswith("chrome-mac")
    ]

    if not mac_dirs:
        raise FileNotFoundError("No mac chromium folder found")

    chromium_path = (
        mac_dirs[0]
        / "Google Chrome for Testing.app"
        / "Contents"
        / "MacOS"
        / "Google Chrome for Testing"
    )






print(f"Using Chromium: {chromium_path}")

# =====================================================
# PLAYWRIGHT
# =====================================================

script_start = time.time()

with sync_playwright() as p:

    context = p.chromium.launch_persistent_context(
        user_data_dir=USER_DATA_DIR,
        executable_path=chromium_path,
        headless=HEADLESS,
        accept_downloads=True
    )


    page = context.new_page()

    for index, record in enumerate(records, start=1):

        row_number = record["row"]

        url = record["url"]

        save_directory = record["directory"]

        print("\n================================================")

        print(f"Processing {index} of {len(records)}")

        record_start = time.time()

        try:

            # =====================================================
            # CREATE DIRECTORY
            # =====================================================

            if not os.path.exists(save_directory):

                print(f"Creating directory: {save_directory}")

                os.makedirs(save_directory, exist_ok=True)

            # =====================================================
            # PAGE LOAD TIMER
            # =====================================================

            page_start = time.time()

            # Open page
            page.goto(url, timeout=120000)

            # Wait for page load
            page.wait_for_load_state("networkidle")

            page_elapsed = time.time() - page_start

            print(
                f"Page Load Time: "
                f"{format_elapsed(page_elapsed)}"
            )

            # Small delay for stability
            page.wait_for_timeout(3000)

            # =====================================================
            # DOWNLOAD
            # =====================================================

            with page.expect_download(timeout=120000) as download_info:

                page.get_by_role(
                    "link",
                    name="DOWNLOAD"
                ).click()

            download = download_info.value

            # Get filename
            filename = download.suggested_filename

            print(f"Downloading: {filename}")

            # Build full save path
            save_path = os.path.join(
                save_directory,
                filename
            )

            # Save file
            download.save_as(save_path)

            print(f"Saved to: {save_directory}")

            print(os.getcwd())

            if platform.system() == "Windows":
                from pathlib import Path

                print(f"Running on Windows: {UNZIP_MOVE_BATCH}")
                subprocess.run([UNZIP_MOVE_BATCH, save_directory], check=True, shell=True)
            else:
                print(f"Running on {platform.system()}")
                subprocess.run([UNZIP_MOVE_BATCH, save_directory], check=True)

            # =====================================================
            # WRITE FILENAME TO COLUMN
            # =====================================================
            ws[f"{FILENAME_COLUMN}{row_number}"] = filename

            # =====================================================
            # WRITE CASE STATIUS TO "DOWNLOADED"
            # =====================================================
            ws[f"{CASE_STATUS}{row_number}"] = "DOWNLOADED"

           # =====================================================
           # Write completion timestamp
           # =====================================================
            timestamp_cell = ws[f"{COMPLETED_COLUMN}{row_number}"]
            timestamp_cell.value = datetime.now()
            timestamp_cell.number_format = "mm/dd/yyyy"

            # Save workbook immediately3
            wb.save(EXCEL_FILE)


            print(f"Updated Excel row {row_number}")

            # =====================================================
            # RECORD TIMER
            # =====================================================

            record_elapsed = time.time() - record_start

            print(
                f"Record Time: "
                f"{format_elapsed(record_elapsed)}"
            )

        except Exception as e:

            record_elapsed = time.time() - record_start

            print("ERROR:")

            print(e)

            print(
                f"Failed After: "
                f"{format_elapsed(record_elapsed)}"
            )

# =====================================================
# FINAL SAVE
# =====================================================

wb.save(EXCEL_FILE)

total_elapsed = time.time() - script_start

print("\nAll downloads complete.")

print(
    f"\nTotal Runtime: "
    f"{format_elapsed(total_elapsed)}"
)

#input("\nPress ENTER to close...")